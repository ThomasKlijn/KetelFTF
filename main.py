from fastapi import FastAPI, File, UploadFile
from fastapi.responses import StreamingResponse, FileResponse
import pandas as pd
import numpy as np
import joblib
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill
import uvicorn

app = FastAPI()

@app.get("/")
async def read_root():
    return FileResponse("index.html")

# Modellen en vectorizers laden
ketel_model = joblib.load("ketel_model_Vfinal.joblib")
ketel_vec = joblib.load("ketel_vectorizer_Vfinal.joblib")
ftf_model = joblib.load("ftf_model_Vfinal.joblib")
ftf_vec = joblib.load("ftf_vectorizer_Vfinal.joblib")

# Keywordlijsten Ketel gerelateerd (lowercase keys voor case-insensitive matching)
keywords_kg = {
    "flowsensor vervangen": "ja",
    "aansturing ivm 9u": "ja",
    "aansturing nefit": "ja",
    "aansturing ivm 9p": "ja",
    "lek hydroblok": "ja",
    "hydroblok": "ja",
    "3wk": "ja",
    "3WK": "ja",
    "branderautomaat": "ja",
    "defecte kim": "ja",
    "platenwisselaar vervangen": "ja",
    "wisselaar vervangen": "ja",
    "het hydroblok was lek": "ja",
    "rga beugelen": "nee",
    "rga gebeugeld": "nee",
    "druksensor vervangen": "ja",
    "condensafvoer": "nee",
    "hoofdprint vervangen": "ja",
    "driewegklep": "ja",
    "drie weg klep": "ja",
    "kaart in de bus": "nee",
    "Kaart in de bus": "nee",
    "3wegklep": "ja",
    "3weg klep": "ja",
    "wisselaar": "ja",
    "stromingssensor": "ja",
    "hoofdprint": "ja",
    "drieweg klep": "ja",
    "ontstek pen vv": "ja",
    "thermostaat van de klant is niet goed": "nee",
    "wartel bij de pomp": "ja",
    "ontsteekpen vervangen": "ja",
    "pomp vervangen": "ja",
    "ketel vervangen": "ja",
    "pomp": "ja",
    "condensafvoer herstellen.": "nee",
    "nieuwe afvoer maken": "nee",
    "flowsensor vv": "ja",
    "sensorflow vervangen": "ja",
    "sensorflow vv": "ja",
    "expansievat": "nee",
    "vloerverwarming": "nee",
    "slang op verdeler": "nee",
    "expansievat vervangen": "nee",
    "rga aanpassen": "nee",
    "warmtewisselaar": "ja",
    "ltv(luchttoevoer) vervangen": "nee",
    "verroest van binnen": "ja",
    "batterijen vervangen": "nee",
    "wisselaar is lek": "ja",
    "ketel bij gevuld": "nee",
    "pomp zat vast": "ja",
    "geen gas toevoer": "nee",
    "batterij van de koolmonoxide melder": "nee",
    "vaillant pomp aansluitkabel": "ja",
    "vaillant druksensor": "ja",
    "thermostaat": "nee",
    "vulkraan vervangen": "nee",
    "stekkertje ontsteekkabel": "nee",
    "rga herstellen": "nee",
    "rga vervangen en gebeugeld": "nee",
    "ltv (luchttoevoer) gebeugeld": "nee",
    "warmtewisselaar vervangen": "ja",
    "sensor flow vervangen": "ja",
    "pomp vv": "ja",
    "pomp verv": "ja",
    "sensor flow": "ja",
    "ketel vv": "ja",
    "flow sensor": "ja",
    "flowsensor": "ja",
    "ketel verv": "ja",
    "ventilator": "ja",
    "fentilator": "ja",
    "wtw": "nee",
    "pakking": "ja",
    "gloeiplug": "ja",
    "3 weg klep": "ja",
    "nieuwe ketel": "ja",
    "sam trechter vervangen": "nee",
    "rga en ltv + dakdoorvoer aanpassen.": "nee",
    "rookgas vervangen": "nee",
    "inspuiter": "ja",
    "bijgevuld": "nee",
    "flowswitch": "ja",
    "druksensor defect": "ja",
    "alle radiatorkranen dicht": "nee",
    "gasblok vervangen": "ja",
    "wasmachinekraan vervangen": "nee",
    "lijkt op print te zijn.": "ja",
    "overstort druppelde": "nee",
    "overstort en wasmachine kraan": "nee",
    "overstort vervangen": "nee",
    "uitgevoerde werkzaamheden: ketel afgekeurd": "ja",
    "spoed - ketel afgekeurd - nog vervangen": "ja",
    "ketel afgekeurd - nog vervangen": "ja",
    "thermostaat vervangen": "nee",
    "afuizing badkamer": "nee",
    "kamerthermostaat vervangen": "nee"
}

# Rule-based FTF keywords
positive_keywords = [
    "vervangen", "gerepareerd", "hersteld", "reset", "opgelost",
    "functioneert", "controle uitgevoerd", "storingscode verwijderd"
]
negative_keywords = [
    "afspraak", "moet worden nagekeken", "onderdeel besteld", 
    "kan niet oplossen", "opvolging", "storingscode blijft", "niet gelukt"
]

def keyword_based_ftf(text: str):
    text_lower = text.lower()
    if any(k in text_lower for k in positive_keywords):
        return "FTF"  # Opgelost = first time fix
    if any(k in text_lower for k in negative_keywords):
        return "1"    # Niet opgelost = storing indicator
    return None

def apply_keywords_per_column(df, keywords_dict, columns):
    df["Ketel gerelateerd_keyword"] = np.nan
    df["Ketel gerelateerd_keyword"] = df["Ketel gerelateerd_keyword"].astype(object)  # fix dtype warning
    for col in columns:
        text_lower = df[col].fillna("").str.lower()
        for kw, val in keywords_dict.items():
            mask = text_lower.str.contains(kw, na=False, regex=False)
            if val == "ja":
                df.loc[mask, "Ketel gerelateerd_keyword"] = "ja"
            else:
                df.loc[mask & df["Ketel gerelateerd_keyword"].isna(), "Ketel gerelateerd_keyword"] = "nee"
    return df

def create_address(df):
    """Combineer adres kolommen tot één adres veld"""
    address_cols = ["Werkbon bezoekadres straat", "Werkbon bezoekadres huisnummer", "toevoeging"]
    
    # Controleer welke kolommen bestaan
    existing_cols = [col for col in address_cols if col in df.columns]
    
    if existing_cols:
        # Combineer bestaande kolommen, vul lege waarden met lege string
        df["Adres"] = df[existing_cols].fillna("").astype(str).agg(" ".join, axis=1).str.strip()
        # Vervang meerdere spaties door één spatie
        df["Adres"] = df["Adres"].str.replace(r'\s+', ' ', regex=True)
    else:
        # Als geen adres kolommen bestaan, maak lege adres kolom
        df["Adres"] = ""
    
    return df

def compute_visit_patterns(df):
    """
    Compute address visit counts and detect multi-visit scenarios.
    Include fallback heuristics for missing address data.
    
    Returns:
    - address_counts: Series with visit counts per address
    - multi_visit_indices: set of row indices that are part of multi-visit scenarios
    - single_visit_indices: set of row indices that are genuinely single visits
    """
    address_counts = pd.Series(dtype=int)
    multi_visit_indices = set()
    single_visit_indices = set()
    
    # Primary method: gebruik Adres kolom als beschikbaar en niet leeg
    if "Adres" in df.columns:
        # Filter out empty addresses
        valid_address_mask = df["Adres"].fillna("").str.strip().ne("")
        
        if valid_address_mask.any():
            # Compute counts voor geldige adressen
            address_counts = df.loc[valid_address_mask, "Adres"].value_counts()
            
            # Identificeer multi-visit addresses
            multi_addresses = address_counts[address_counts >= 2].index.tolist()
            
            for address in multi_addresses:
                if address.strip():  # Extra veiligheidscheck
                    mask = df["Adres"] == address
                    multi_visit_indices.update(df.loc[mask].index.tolist())
            
            # Single visit addresses (met geldig adres)
            single_addresses = address_counts[address_counts == 1].index.tolist()
            for address in single_addresses:
                if address.strip():
                    mask = df["Adres"] == address
                    single_visit_indices.update(df.loc[mask].index.tolist())
            
            print(f"Address method: {len(multi_addresses)} multi-visit addresses, {len(single_addresses)} single-visit addresses")
    
    # Fallback heuristics voor rijen zonder geldig adres
    remaining_indices = set(df.index) - multi_visit_indices - single_visit_indices
    
    if remaining_indices:
        print(f"Applying fallback heuristics for {len(remaining_indices)} rows without valid address")
        
        # Fallback 1: Werkbon follow-up chains (als kolommen beschikbaar zijn)
        if "Werkbon nummer" in df.columns and "Werkbon is vervolg van" in df.columns:
            # Build follow-up map
            follow_up_map = {}
            wb_to_index = {}
            
            for idx in remaining_indices:
                row = df.loc[idx]
                wb_num = row["Werkbon nummer"]
                wb_to_index[wb_num] = idx
                
                follow_up = row["Werkbon is vervolg van"]
                if pd.notna(follow_up) and str(follow_up).strip():
                    if str(follow_up).startswith('WB'):
                        follow_up_map[wb_num] = follow_up
            
            # Identify chains
            def find_all_in_chain(wb):
                """Find all workbons in the same chain"""
                chain = set()
                # Find all ancestors
                current = wb
                visited = set()
                while current in follow_up_map and current not in visited:
                    visited.add(current)
                    current = follow_up_map[current]
                
                # Current is now the root, collect all descendants
                chain.add(current)
                changed = True
                while changed:
                    changed = False
                    for child, parent in follow_up_map.items():
                        if parent in chain and child not in chain:
                            chain.add(child)
                            changed = True
                
                return chain
            
            processed_wbs = set()
            for wb in follow_up_map.keys():
                if wb not in processed_wbs:
                    chain_wbs = find_all_in_chain(wb)
                    processed_wbs.update(chain_wbs)
                    
                    if len(chain_wbs) >= 2:
                        # Multi-visit scenario
                        chain_indices = [wb_to_index[wb] for wb in chain_wbs if wb in wb_to_index and wb_to_index[wb] in remaining_indices]
                        multi_visit_indices.update(chain_indices)
                        remaining_indices -= set(chain_indices)
            
            print(f"Fallback 1 (follow-up chains): identified {len(multi_visit_indices) - len([i for i in multi_visit_indices if i in df.index and df.loc[i, 'Adres'].strip()])} additional multi-visit rows")
        
        # Fallback 2: Frequent text patterns (conservative heuristic)
        if remaining_indices and "combined_text" in df.columns:
            # Extract key text features voor duplicate detection
            remaining_df = df.loc[list(remaining_indices)].copy()
            
            # Simplistic text similarity based op keywords
            if len(remaining_df) > 1:
                # Create text signatures (first few significant words)
                def create_text_signature(text):
                    if pd.isna(text) or not str(text).strip():
                        return ""
                    words = str(text).lower().split()
                    # Take first 3-5 significant words (excluding common stop words)
                    stop_words = {'de', 'het', 'en', 'van', 'in', 'op', 'met', 'is', 'een', 'voor', 'aan', 'na', 'bij'}
                    sig_words = [w for w in words[:10] if len(w) > 2 and w not in stop_words][:5]
                    return " ".join(sig_words)
                
                remaining_df["text_signature"] = remaining_df["combined_text"].apply(create_text_signature)
                
                # Find signature duplicates (very conservative threshold)
                sig_counts = remaining_df["text_signature"].value_counts()
                dup_sigs = sig_counts[sig_counts >= 2].index.tolist()
                
                for sig in dup_sigs:
                    if sig.strip():  # Ignore empty signatures
                        sig_mask = remaining_df["text_signature"] == sig
                        sig_indices = remaining_df.loc[sig_mask].index.tolist()
                        if len(sig_indices) >= 2:
                            multi_visit_indices.update(sig_indices)
                            remaining_indices -= set(sig_indices)
                
                print(f"Fallback 2 (text patterns): identified {len(dup_sigs)} potential duplicate patterns")
        
        # All remaining indices are considered single-visit
        single_visit_indices.update(remaining_indices)
    
    print(f"Final classification: {len(multi_visit_indices)} multi-visit, {len(single_visit_indices)} single-visit rows")
    
    return address_counts, multi_visit_indices, single_visit_indices

def apply_duplicate_address_rule(df):
    """Regel: Bij dubbele adressen, sorteer op datum en stel FTF in"""
    if "Adres" not in df.columns or "Uitvoerdatum" not in df.columns:
        return df, set()
    
    # Track alle rijen die door deze regel worden aangepast
    business_rule_indices = set()
    
    # Zoek adressen die 2+ keer voorkomen
    address_counts = df["Adres"].value_counts()
    duplicate_addresses = address_counts[address_counts >= 2].index.tolist()
    
    for address in duplicate_addresses:
        if address.strip():  # Negeer lege adressen
            # Filter rijen met dit adres
            mask = df["Adres"] == address
            address_rows = df.loc[mask].copy()
            
            # Converteer Uitvoerdatum naar datetime voor juiste sortering
            try:
                address_rows["Uitvoerdatum_dt"] = pd.to_datetime(address_rows["Uitvoerdatum"], errors='coerce')
                # Sorteer op datetime (oudste eerst)
                address_rows = address_rows.sort_values("Uitvoerdatum_dt")
            except:
                # Fallback naar string sortering als datetime conversie mislukt
                address_rows = address_rows.sort_values("Uitvoerdatum")
            
            address_indices = address_rows.index.tolist()
            
            # Stel FTF in: oudste = "NFT", nieuwste = "1", rest = "" (leeg)
            if len(address_indices) >= 2:
                df.loc[address_indices[0], "FTF"] = "NFT"  # Oudste (geen FTF mogelijk)
                df.loc[address_indices[-1], "FTF"] = "1"   # Nieuwste (storing indicator)
                
                # Alle tussenliggende rijen leeg maken
                for idx in address_indices[1:-1]:
                    df.loc[idx, "FTF"] = ""
                
                # Track ALLE aangepaste rijen (inclusief degene die leeg worden)
                business_rule_indices.update(address_indices)
    
    return df, business_rule_indices

def apply_werkbon_follow_up_rule(df):
    """Regel: Bij werkbon follow-ups, bouw complete chains en sorteer op datum
    KRITISCH: Respecteer bestaande FTF toewijzingen van address rules!"""
    if "Werkbon nummer" not in df.columns or "Werkbon is vervolg van" not in df.columns or "Uitvoerdatum" not in df.columns:
        return df, set()
    
    # Track alle rijen die door deze regel worden aangepast
    business_rule_indices = set()
    
    # Bouw een graph van alle follow-up relaties
    follow_up_map = {}  # child -> parent mapping
    wb_to_index = {}    # werkbon -> df index mapping
    
    # Vul mappings
    for idx, row in df.iterrows():
        wb_num = row["Werkbon nummer"]
        wb_to_index[wb_num] = idx
        
        follow_up = row["Werkbon is vervolg van"]
        if pd.notna(follow_up) and str(follow_up).strip():
            # Check voor WB format
            if str(follow_up).startswith('WB'):
                follow_up_map[wb_num] = follow_up
    
    # Vind alle complete chains
    processed_workbons = set()
    
    def find_chain_root(wb):
        """Vind de root (oudste) werkbon in een chain"""
        current = wb
        visited = set()
        while current in follow_up_map and current not in visited:
            visited.add(current)
            current = follow_up_map[current]
        return current
    
    def build_complete_chain(root_wb):
        """Bouw complete chain vanaf root"""
        chain = [root_wb]
        # Vind alle children recursief
        changed = True
        while changed:
            changed = False
            for child, parent in follow_up_map.items():
                if parent in chain and child not in chain:
                    chain.append(child)
                    changed = True
        return chain
    
    # Process elke chain
    for wb in follow_up_map.keys():
        if wb not in processed_workbons:
            # Vind de root van deze chain
            root_wb = find_chain_root(wb)
            
            # Bouw complete chain
            chain_workbons = build_complete_chain(root_wb)
            
            # Markeer als processed
            processed_workbons.update(chain_workbons)
            
            # Als we een chain van 2+ hebben, check of we deze kunnen behandelen
            if len(chain_workbons) >= 2:
                # Verzamel indices voor alle workbons in chain
                chain_indices = []
                
                for chain_wb in chain_workbons:
                    if chain_wb in wb_to_index:
                        idx = wb_to_index[chain_wb]
                        chain_indices.append(idx)
                
                if len(chain_indices) >= 2:
                    # KRITISCH: Check of er al address rule toewijzingen zijn
                    existing_ftf_assignments = df.loc[chain_indices, "FTF"].fillna("").ne("").any()
                    
                    if existing_ftf_assignments:
                        # Address rules hebben al toegewezen - respecteer deze volledig
                        # Track deze rijen als aangepast (zodat ML/keywords ze niet overschrijven)
                        business_rule_indices.update(chain_indices)
                        print(f"Follow-up rule respecteert bestaande address rule assignments voor chain: {chain_workbons}")
                    else:
                        # Geen conflicten - pas normale follow-up logica toe
                        # Maak DataFrame van chain rows voor sortering
                        chain_df = df.loc[chain_indices].copy()
                        
                        # Converteer Uitvoerdatum naar datetime voor juiste sortering
                        try:
                            chain_df["Uitvoerdatum_dt"] = pd.to_datetime(chain_df["Uitvoerdatum"], errors='coerce')
                            # Sorteer op datetime (oudste eerst)
                            chain_df = chain_df.sort_values("Uitvoerdatum_dt")
                        except:
                            # Fallback naar string sortering als datetime conversie mislukt
                            chain_df = chain_df.sort_values("Uitvoerdatum")
                        
                        sorted_indices = chain_df.index.tolist()
                        
                        # Stel FTF in: oudste = "NFT", nieuwste = "1", rest = "" (leeg)
                        df.loc[sorted_indices[0], "FTF"] = "NFT"  # Oudste (geen FTF mogelijk)
                        df.loc[sorted_indices[-1], "FTF"] = "1"   # Nieuwste (storing indicator)
                        
                        # Alle tussenliggende rijen leeg maken
                        for sidx in sorted_indices[1:-1]:
                            df.loc[sidx, "FTF"] = ""
                        
                        # Track ALLE aangepaste rijen (inclusief degene die leeg worden)
                        business_rule_indices.update(sorted_indices)
    
    return df, business_rule_indices

@app.post("/process_excel/")
async def process_excel(file: UploadFile = File(...)):
    contents = await file.read()
    df_prod = pd.read_excel(BytesIO(contents), sheet_name="Leeg")

    alle_kolommen = df_prod.columns.tolist()
    potentiele_oplossingskolommen = ["Oplossingen"] + [col for col in alle_kolommen if col.startswith("Unnamed:")]
    oplossingskolommen = [col for col in potentiele_oplossingskolommen if col in alle_kolommen]
    
    basis_tekstkolommen = [
        "Werkbeschrijving", "Werkbon is vervolg van",
        "Werkbon nummer", "Uitvoerdatum", "Object referentie", "Installatie apparaat omschrijving"
    ]
    tekstkolommen = [col for col in basis_tekstkolommen if col in alle_kolommen]

    if oplossingskolommen:
        df_prod[oplossingskolommen] = df_prod[oplossingskolommen].fillna("")
        df_prod["Oplossingen_samengevoegd"] = df_prod[oplossingskolommen].astype(str).agg(" ".join, axis=1)
    else:
        df_prod["Oplossingen_samengevoegd"] = ""

    tekstkolommen.append("Oplossingen_samengevoegd")
    df_prod[tekstkolommen] = df_prod[tekstkolommen].fillna("")
    df_prod["combined_text"] = df_prod[tekstkolommen].apply(lambda r: " ".join([str(x) for x in r]), axis=1)
    df_prod["heeft_vervolg"] = df_prod["Werkbon is vervolg van"].apply(lambda x: int(bool(str(x).strip())))
    df_prod["tekstlengte"] = df_prod["combined_text"].apply(len)
    df_prod["woordenaantal"] = df_prod["combined_text"].apply(lambda x: len(x.split()))
    df_prod["contains_onderdeel"] = df_prod["combined_text"].str.contains("onderdeel|vervangen|vervang", case=False).astype(int)
    df_prod["contains_reset"] = df_prod["combined_text"].str.contains("reset|herstart", case=False).astype(int)
    df_prod["contains_advies"] = df_prod["combined_text"].str.contains("advies|aanbeveling", case=False).astype(int)

    # ---- Keyword matching Ketel gerelateerd, per kolom met prioriteit ----
    keyword_check_columns = ["Werkbeschrijving", "Oplossingen_samengevoegd"]
    df_prod = apply_keywords_per_column(df_prod, keywords_kg, keyword_check_columns)

    # Vul kolom Ketel gerelateerd vanuit keywords (prioriteit)
    mask_keyword = df_prod["Ketel gerelateerd_keyword"].notnull()
    df_prod["Ketel gerelateerd"] = df_prod["Ketel gerelateerd"].astype(object)  # fix dtype warning
    df_prod.loc[mask_keyword, "Ketel gerelateerd"] = df_prod.loc[mask_keyword, "Ketel gerelateerd_keyword"]

    # Model voorspelling Ketel gerelateerd alleen toepassen op rijen zonder keyword invulling
    mask_ml = df_prod["Ketel gerelateerd"].isnull() | (df_prod["Ketel gerelateerd"] == "")
    if mask_ml.any():
        X_text_k = df_prod.loc[mask_ml, "combined_text"]
        X_vec_k = ketel_vec.transform(X_text_k).toarray()
        extra_k = df_prod.loc[mask_ml, ["heeft_vervolg", "tekstlengte", "woordenaantal", "contains_onderdeel", "contains_reset", "contains_advies"]].values
        X_comb_k = np.hstack((X_vec_k, extra_k))

        y_proba_k = ketel_model.predict_proba(X_comb_k)
        y_pred_k = ketel_model.predict(X_comb_k)

        df_prod.loc[mask_ml, "Ketel gerelateerd"] = np.where(y_proba_k.max(axis=1) > 0.7, np.where(y_pred_k == 1, "ja", "nee"), "")

    df_prod["Ketel zekerheid"] = 0.0
    if mask_ml.any():
        df_prod.loc[mask_ml, "Ketel zekerheid"] = y_proba_k.max(axis=1)
    df_prod.loc[mask_keyword, "Ketel zekerheid"] = 1.0  # Keywords zijn zeker

    # FTF wordt niet meer voorspeld in deze stap

    # Kolommen die NIET in output mogen
    exclude_cols = [
        "Unnamed: 18", "Unnamed: 19", "Unnamed: 20", "Unnamed: 21", "Unnamed: 22", "Unnamed: 23",
        "Oplossingen_samengevoegd", "combined_text", "heeft_vervolg", "tekstlengte", "woordenaantal",
        "contains_onderdeel", "contains_reset", "contains_advies", "Ketel zekerheid",
        "Ketel gerelateerd_keyword"
    ]

    output_cols = [col for col in df_prod.columns if col not in exclude_cols]

    # Kolommen die wél in output blijven maar zonder kolomnaam in Excel
    cols_with_empty_name = [
        "Werkbon is vervolg van", "Werkbon nummer", "Uitvoerdatum", "Object referentie",
        "Installatie apparaat omschrijving"
    ]

    output_col_names = []
    for col in output_cols:
        if col in cols_with_empty_name:
            output_col_names.append("")
        else:
            output_col_names.append(col)

    # Output Excel met kleurcodering
    wb = openpyxl.Workbook()
    ws = wb.active

    # Schrijf header (kolomnamen)
    for col_idx, col_name in enumerate(output_col_names, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    geel = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    oranje = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    for row_idx, row in df_prod.iterrows():
        excel_row = row_idx + 2
        for col_idx, col_name in enumerate(output_cols, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=row[col_name])

            # Kleur Ketel gerelateerd geel bij zekerheid >= 0.7, oranje bij 0 < zekerheid < 0.7
            if col_name == "Ketel gerelateerd":
                if row["Ketel zekerheid"] >= 0.7:
                    cell.fill = geel
                elif 0 < row["Ketel zekerheid"] < 0.7:
                    cell.fill = oranje

            # FTF kleurcodering wordt niet meer toegepast in deze stap

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=voorspeld_{file.filename}"}
    )

@app.post("/split_output/")
async def split_output(file: UploadFile = File(...)):
    contents = await file.read()
    df = pd.read_excel(BytesIO(contents))

    df_niet_ketel = df[df["Ketel gerelateerd"] == "nee"]
    df_ketel_gerelateerd = df[df["Ketel gerelateerd"] == "ja"]

    stream = BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        df_niet_ketel.to_excel(writer, index=False, sheet_name="Niet ketel gerelateerd")
        df_ketel_gerelateerd.to_excel(writer, index=False, sheet_name="Ketel gerelateerd")
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gesplitst_{file.filename}"}
    )

@app.post("/predict_ftf/")
async def predict_ftf(file: UploadFile = File(...)):
    contents = await file.read()
    
    # Probeer alle werkbladen te lezen
    try:
        all_sheets = pd.read_excel(BytesIO(contents), sheet_name=None)
    except:
        # Als het lezen mislukt, behandel als single sheet
        df = pd.read_excel(BytesIO(contents))
        all_sheets = {"Sheet1": df}
    
    # Zoek naar het "Ketel gerelateerd" werkblad, anders gebruik het eerste werkblad met 'j' rijen
    target_sheet = None
    target_sheet_name = None
    
    if "Ketel gerelateerd" in all_sheets:
        target_sheet = all_sheets["Ketel gerelateerd"]
        target_sheet_name = "Ketel gerelateerd"
    else:
        # Zoek naar werkblad met "Ketel gerelateerd" = "ja" rijen
        for sheet_name, sheet_df in all_sheets.items():
            if "Ketel gerelateerd" in sheet_df.columns:
                mask_ftf = (sheet_df["Ketel gerelateerd"] == "ja")
                if mask_ftf.any():
                    target_sheet = sheet_df
                    target_sheet_name = sheet_name
                    break
    
    if target_sheet is None:
        # Geen geschikt werkblad gevonden
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Geen werkblad gevonden met 'Ketel gerelateerd' = 'ja' rijen. Upload het gesplitste bestand uit stap 2.")
    
    df = target_sheet.copy()
    
    # Alleen rijen waar "Ketel gerelateerd" = "ja" behandelen voor FTF voorspelling
    mask_ftf = (df["Ketel gerelateerd"] == "ja")
    df_ftf = df[mask_ftf].copy()
    
    if len(df_ftf) == 0:
        # Geen ketel gerelateerde rijen gevonden in dit werkblad
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Geen rijen met 'Ketel gerelateerd' = 'ja' gevonden in het gekozen werkblad.")
    
    # Bereid tekst features voor
    alle_kolommen = df_ftf.columns.tolist()
    
    # *** DETERMINISTISCHE & UNIEKE MAPPING: Detecteer kolommen op basis van inhoud ***
    # Als Excel geen header heeft, worden kolommen "Unnamed: X", etc.
    column_mapping = {}
    unnamed_cols = [col for col in alle_kolommen if str(col).startswith("Unnamed:")]
    
    if unnamed_cols:
        print(f"Gevonden unnamed kolommen: {unnamed_cols}")
        
        # Bereken COVERAGE-SENSITIVE metrics op FULL dataframe (niet alleen df_ftf subset)
        col_metrics = {}
        for col in unnamed_cols:
            # Gebruik volledige dataset voor betrouwbaardere metrics
            full_col_data = df[col].astype(str)
            non_null_mask = df[col].notna() & (df[col].astype(str).str.strip() != "")
            non_null_data = df.loc[non_null_mask, col].astype(str)
            
            # Density = non-null coverage
            density = non_null_mask.sum() / len(df) if len(df) > 0 else 0
            
            if len(non_null_data) == 0:
                col_metrics[col] = {'werkbon_score': 0, 'vervolg_score': 0, 'date_score': 0, 'density': 0}
                continue
                
            # Verbeterde WB pattern matching (handelt variaties af)
            wb_pattern = r'^\s*WB[- ]?\d+'
            wb_matches = non_null_data.str.contains(wb_pattern, na=False, regex=True).sum()
            wb_match_fraction = wb_matches / len(df)  # Fractie van ALLE rijen
            
            # Werkbon score: WB matches gewogen met density (voorkomt sparse high-scoring vervolg cols)
            werkbon_score = wb_match_fraction * density
            
            # Vervolg score: hoge null rate + enkele WB matches (prioriteert sparsity)
            null_fraction = (~non_null_mask).sum() / len(df)
            vervolg_score = null_fraction * (1 if wb_matches > 0 else 0)
            
            # Datum score: datetime conversie success rate op non-null data
            try:
                date_successes = pd.to_datetime(non_null_data, errors='coerce').notna().sum()
                date_score = (date_successes / len(non_null_data)) * density
            except:
                date_score = 0
            
            col_metrics[col] = {
                'werkbon_score': werkbon_score,
                'vervolg_score': vervolg_score, 
                'date_score': date_score,
                'density': density,
                'wb_matches': wb_matches,
                'null_fraction': null_fraction
            }
            
            print(f"Metrics voor {col}: werkbon_score={werkbon_score:.3f}, vervolg_score={vervolg_score:.3f}, "
                  f"date_score={date_score:.3f}, density={density:.3f}")
        
        # ROLE-PRECEDENCE TOEWIJZING: Vervolg eerst, dan Werkbon, dan Datum
        # Dit voorkomt dat sparse vervolg-kolommen verkeerd als werkbon worden geclassificeerd
        used_cols = set()
        assignment_order = [
            ("Werkbon is vervolg van", 'vervolg_score', 0.3),  # Lagere threshold voor vervolg
            ("Werkbon nummer", 'werkbon_score', 0.1),          # Werkbon na vervolg
            ("Uitvoerdatum", 'date_score', 0.3)               # Datum als laatste
        ]
        
        for target_name, score_key, min_threshold in assignment_order:
            # Skip als target al bestaat
            if target_name in alle_kolommen:
                continue
                
            best_col = None
            best_score = 0
            
            # Vind beste kandidaat voor deze role
            for col, metrics in col_metrics.items():
                if col in used_cols:
                    continue
                    
                score = metrics[score_key]
                
                # Voor werkbon: extra check dat het niet beter past als vervolg
                if target_name == "Werkbon nummer" and metrics['vervolg_score'] > metrics['werkbon_score']:
                    continue  # Skip deze kolom, het is waarschijnlijk een vervolg kolom
                
                if score > best_score and score >= min_threshold:
                    best_score = score
                    best_col = col
            
            if best_col:
                column_mapping[best_col] = target_name
                used_cols.add(best_col)
                print(f"Mapped {best_col} -> {target_name} (score: {best_score:.3f}, threshold: {min_threshold})")
            else:
                print(f"Geen geschikte kolom gevonden voor {target_name} (min threshold: {min_threshold})")
        
        # Pas mapping toe
        if column_mapping:
            df_ftf = df_ftf.rename(columns=column_mapping)
            df = df.rename(columns=column_mapping)  # Ook originele df updaten
            alle_kolommen = df_ftf.columns.tolist()
            print(f"Finale mapping: {column_mapping}")
        else:
            print("Geen betrouwbare mapping gevonden (scores te laag)")
    
    # Exclusief gemappte unnamed kolommen van oplossingskolommen (voorkomen text contamination)
    mapped_unnamed_cols = set(column_mapping.keys()) if column_mapping else set()
    potentiele_oplossingskolommen = ["Oplossingen"] + [col for col in alle_kolommen if col.startswith("Unnamed:") and col not in mapped_unnamed_cols]
    oplossingskolommen = [col for col in potentiele_oplossingskolommen if col in alle_kolommen]
    
    basis_tekstkolommen = [
        "Werkbeschrijving", "Werkbon is vervolg van",
        "Werkbon nummer", "Uitvoerdatum", "Object referentie", "Installatie apparaat omschrijving"
    ]
    tekstkolommen = [col for col in basis_tekstkolommen if col in alle_kolommen]

    if oplossingskolommen:
        df_ftf[oplossingskolommen] = df_ftf[oplossingskolommen].fillna("")
        df_ftf["Oplossingen_samengevoegd"] = df_ftf[oplossingskolommen].astype(str).agg(" ".join, axis=1)
    else:
        df_ftf["Oplossingen_samengevoegd"] = ""

    tekstkolommen.append("Oplossingen_samengevoegd")
    df_ftf[tekstkolommen] = df_ftf[tekstkolommen].fillna("")
    df_ftf["combined_text"] = df_ftf[tekstkolommen].apply(lambda r: " ".join([str(x) for x in r]), axis=1)
    
    # Defensieve controle voor "Werkbon is vervolg van" kolom
    if "Werkbon is vervolg van" in df_ftf.columns:
        df_ftf["heeft_vervolg"] = df_ftf["Werkbon is vervolg van"].apply(lambda x: int(bool(str(x).strip())))
    else:
        df_ftf["heeft_vervolg"] = 0  # Geen follow-up informatie beschikbaar
        
    df_ftf["tekstlengte"] = df_ftf["combined_text"].apply(len)
    df_ftf["woordenaantal"] = df_ftf["combined_text"].apply(lambda x: len(x.split()))
    df_ftf["contains_onderdeel"] = df_ftf["combined_text"].str.contains("onderdeel|vervangen|vervang", case=False).astype(int)
    df_ftf["contains_reset"] = df_ftf["combined_text"].str.contains("reset|herstart", case=False).astype(int)
    df_ftf["contains_advies"] = df_ftf["combined_text"].str.contains("advies|aanbeveling", case=False).astype(int)
    
    # ---- NIEUWE BUSINESS REGELS VOOR FTF ----
    
    # Initialiseer FTF kolom als leeg
    df_ftf["FTF"] = ""
    
    # Track alle rijen die door business rules worden aangepast
    all_business_rule_indices = set()
    
    # Regel 1: Maak adres veld
    df_ftf = create_address(df_ftf)
    
    # KRITISCH: Compute visit patterns VOOR business rules om multi-visit scenarios te detecteren
    address_counts, multi_visit_indices, single_visit_indices = compute_visit_patterns(df_ftf)
    
    # Regel 2: Dubbele adressen behandelen (AUTHORITATIEF - finale beslissing)
    df_ftf, address_rule_indices = apply_duplicate_address_rule(df_ftf)
    all_business_rule_indices.update(address_rule_indices)
    
    # Regel 3: Werkbon follow-up chains behandelen (RESPECTEERT address rule assignments)
    df_ftf, followup_rule_indices = apply_werkbon_follow_up_rule(df_ftf)
    all_business_rule_indices.update(followup_rule_indices)
    
    # KRITISCH: ML/keywords ALLEEN voor single-visit addresses die NIET door business rules zijn behandeld
    # Combineer business rule exclusion met single-visit requirement
    mask_eligible_for_ml_keywords = (
        ~df_ftf.index.isin(all_business_rule_indices) &  # Niet door business rules aangepast
        df_ftf.index.isin(single_visit_indices)          # Alleen single-visit scenarios
    )
    
    print(f"ML/Keywords eligible rows: {mask_eligible_for_ml_keywords.sum()} out of {len(df_ftf)} total rows")
    print(f"Excluded by business rules: {len(all_business_rule_indices)}")
    print(f"Excluded by multi-visit: {len(multi_visit_indices)}")
    
    # Rule-based FTF toewijzen (alleen voor eligible single-visit rijen)
    df_ftf["FTF_keyword"] = None
    if mask_eligible_for_ml_keywords.any():
        df_ftf.loc[mask_eligible_for_ml_keywords, "FTF_keyword"] = df_ftf.loc[mask_eligible_for_ml_keywords, "combined_text"].apply(keyword_based_ftf)

    # ML FTF voorspelling voor eligible rijen zonder rule-based FTF 
    mask_ml_ftf = (df_ftf["FTF_keyword"].isnull()) & mask_eligible_for_ml_keywords
    if mask_ml_ftf.any():
        X_text_f = df_ftf.loc[mask_ml_ftf, "combined_text"]
        X_vec_f = ftf_vec.transform(X_text_f).toarray()
        extra_f = df_ftf.loc[mask_ml_ftf, ["heeft_vervolg", "tekstlengte", "woordenaantal", "contains_onderdeel", "contains_reset", "contains_advies"]].values
        X_comb_f = np.hstack((X_vec_f, extra_f))

        y_proba_f = ftf_model.predict_proba(X_comb_f)
        y_pred_f = ftf_model.predict(X_comb_f)

        # Map ML output: "1" -> "FTF", "0" -> "1", anders leeg
        ftf_pred_raw = np.where(y_proba_f.max(axis=1) > 0.7, y_pred_f.astype(str), "")
        ftf_pred = np.where(ftf_pred_raw == "1", "FTF",    # Opgelost = first time fix
                   np.where(ftf_pred_raw == "0", "1", ""))  # Niet opgelost = storing indicator
        df_ftf.loc[mask_ml_ftf, "FTF_keyword"] = ftf_pred

    # Vul FTF kolom alleen voor eligible single-visit rijen
    # Business rule waarden blijven intact (inclusief lege waarden van multi-visit scenarios)
    if mask_eligible_for_ml_keywords.any():
        keyword_results = df_ftf.loc[mask_eligible_for_ml_keywords, "FTF_keyword"].fillna("")
        df_ftf.loc[mask_eligible_for_ml_keywords, "FTF"] = keyword_results
    
    # Update originele dataframe
    df.loc[df_ftf.index, "FTF"] = df_ftf["FTF"]

    # FTF zekerheid alleen voor ML voorspellingen
    df_ftf["FTF zekerheid"] = 0.0
    if mask_ml_ftf.any():
        df_ftf.loc[mask_ml_ftf, "FTF zekerheid"] = y_proba_f.max(axis=1)
    df.loc[df_ftf.index, "FTF zekerheid"] = df_ftf["FTF zekerheid"]
    
    # Update het originele werkblad in all_sheets
    all_sheets[target_sheet_name] = df
    
    # Kolommen die NIET in output mogen
    exclude_cols = [
        "Oplossingen_samengevoegd", "combined_text", "heeft_vervolg", "tekstlengte", "woordenaantal",
        "contains_onderdeel", "contains_reset", "contains_advies", "FTF zekerheid", "Adres"
    ]

    # Maak output Excel bestand met alle werkbladen
    stream = BytesIO()
    with pd.ExcelWriter(stream, engine="openpyxl") as writer:
        for sheet_name, sheet_df in all_sheets.items():
            # Filter kolommen voor output
            output_cols = [col for col in sheet_df.columns if col not in exclude_cols]
            output_df = sheet_df[output_cols].copy()
            
            # Schrijf naar Excel
            output_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Alleen kleurcodering toepassen op het werkblad met FTF voorspellingen
            if sheet_name == target_sheet_name:
                wb = writer.book
                ws = wb[sheet_name]
                
                geel = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                oranje = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                
                # Zoek FTF kolom index
                ftf_col_idx = None
                if "FTF" in output_cols:
                    ftf_col_idx = output_cols.index("FTF") + 1  # Excel is 1-indexed
                
                if ftf_col_idx:
                    for row_idx, row in output_df.iterrows():
                        excel_row = row_idx + 2  # +2 omdat rij 1 header is en pandas is 0-indexed
                        ftf_certainty = sheet_df.loc[row_idx, "FTF zekerheid"] if "FTF zekerheid" in sheet_df.columns else 0
                        ftf_value = str(row.get("FTF", ""))
                        
                        if ftf_value == "1" and ftf_certainty >= 0.6:
                            ws.cell(row=excel_row, column=ftf_col_idx).fill = geel
                        elif 0 < ftf_certainty < 0.6:
                            ws.cell(row=excel_row, column=ftf_col_idx).fill = oranje

    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ftf_voorspeld_{file.filename}"}
    )

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))  # Render uses PORT env var, Replit uses 5000
    uvicorn.run(app, host="0.0.0.0", port=port)
